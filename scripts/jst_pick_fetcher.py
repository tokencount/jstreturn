"""JST fetcher: 拣货仓位 / 主仓 / 排除配件.

Menu path:
  仓储 → 仓位库存
  仓位类型 = 拣货仓位
  仓库区域 = 主仓
  排除标签 = 配件
  → 查询 → 导出

Accounts: 66 / 88 / 99 — passwords read from Mac Keychain via keychain helper.
No passwords stored in source / logs / backups.
"""
from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import re
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from playwright.async_api import async_playwright

log = logging.getLogger("jstreturn.jst_pick")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

JST_BASE_URL = os.environ.get("JST_URL", "https://asia.jsterp.com/Account/Login/index")
HEADLESS = os.environ.get("JST_HEADLESS", "true").lower() != "false"

JST_ACCOUNTS = [
    {"id": "66", "email": "jstpush66@waparcel.com"},
    {"id": "88", "email": "jstpush88@waparcel.com"},
    {"id": "99", "email": "jstpush99@waparcel.com"},
]
if os.environ.get("JST_TEST_ACCOUNT"):
    JST_ACCOUNTS = [row for row in JST_ACCOUNTS if row["id"] == os.environ["JST_TEST_ACCOUNT"]]

DOWNLOAD_WAIT = 60  # seconds
CHROME = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"


# ---------------------------------------------------------------------------
# Keychain helper
# ---------------------------------------------------------------------------

def _keychain_password(service: str, account: str) -> str:
    """Read password from Mac Keychain via security CLI."""
    for cmd in [
        ["security", "find-generic-password", "-s", service, "-a", account, "-w"],
        ["security", "find-internet-password", "-s", service, "-a", account, "-w"],
    ]:
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
        except Exception:
            pass
    return ""


def _keychain_json(service: str) -> dict:
    result = subprocess.run(
        ["/usr/bin/security", "find-generic-password", "-s", service, "-w"],
        capture_output=True, text=True, timeout=10,
    )
    if result.returncode != 0:
        return {}
    try:
        return json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        return {}


# ---------------------------------------------------------------------------
# Login helpers
# ---------------------------------------------------------------------------

async def _login(page, email: str, account_id: str):
    """Navigate to JST and log in. Keychain service = 'JST-{account_id}'."""
    await page.goto(JST_BASE_URL, wait_until="networkidle", timeout=30_000)

    # Keychain service name for this account
    credential = _keychain_json(f"jstreturn-jsterp-{account_id}") or _keychain_json(f"openclaw-jst-{account_id}-runtime")
    username = credential.get("email") or credential.get("username") or credential.get("account") or email
    password = credential.get("password", "")

    if not password:
        raise RuntimeError(
            f"protected JST credentials unavailable for account {account_id}"
        )

    # Check if already logged in
    if "login" not in page.url.lower():
        log.info("JST %s: already logged in at %s", account_id, page.url)
        return

    log.info("JST %s: performing login as %s", account_id, email)
    await page.fill('input[type="text"]', username)
    await page.fill('input[type="password"]', password)
    button = page.get_by_text("立即登录", exact=True).first
    if not await button.is_visible():
        button = page.locator('button[type="submit"]').first
    await button.click()
    await page.wait_for_timeout(2500)
    if await page.locator('input[type="password"]').is_visible():
        raise RuntimeError(f"JST {account_id}: login did not complete")
    log.info("JST %s: login successful", account_id)


# ---------------------------------------------------------------------------
# Navigation
# ---------------------------------------------------------------------------

async def _open_warehouse_inventory(page):
    """仓储 → 仓位库存."""
    # Click 仓储 menu (try multiple selector patterns)
    for selector in [
        'li:has-text("仓储")',
        'a:has-text("仓储")',
        '[class*="menu"] li:has-text("仓储")',
        'text="仓储"',
    ]:
        el = page.locator(selector)
        if await el.count() > 0:
            await el.first.click()
            await page.wait_for_timeout(500)
            break

    # Click 仓位库存 sub-menu
    for selector in [
        'text="仓位库存"',
        'a:has-text("仓位库存")',
        'li:has-text("仓位库存")',
    ]:
        el = page.locator(selector)
        if await el.count() > 0:
            await el.first.click()
            await page.wait_for_timeout(1000)
            break

    # JST preserves stale tabs; force the main business iframe to the
    # verified PackItem route after the menu click.
    changed = await page.evaluate(r"""() => {
      const frames = [...document.querySelectorAll('iframe')].filter(el =>
        !/MICHAT/i.test(el.id || '') && !/onlinecs\.jsterp\.com/.test(el.src || ''));
      const target = frames.find(el => el.getBoundingClientRect().width > 200) || frames[0];
      if (!target) return false;
      target.src = 'https://asia.jsterp.com/Wms/Pack/PackItem?__allsku=' + Date.now();
      return true;
    }""")
    if not changed:
        raise RuntimeError("JST main inventory iframe not found")
    await page.wait_for_timeout(5000)
    log.info("JST: opened 仓位库存 PackItem page")


# ---------------------------------------------------------------------------
# Filters
# ---------------------------------------------------------------------------

SEL_LOCATION_TYPE  = 'input[placeholder="仓位类型"]'
SEL_WAREHOUSE_ZONE = 'input[placeholder="仓库区域"]'
SEL_EXCLUDE_TAG   = 'input[placeholder="排除标签"]'
SEL_QUERY_BTN     = 'button:has-text("查询"), button:has-text("查询")'
SEL_EXPORT_BTN    = 'button:has-text("导出"), .el-button--primary:has-text("导出")'


async def _apply_filters(page,
                         location_type: str = "拣货仓位",
                         warehouse_zone: str = "主仓",
                         exclude_tag: str = "配件"):
    """Fill the three filter fields and click 查询."""
    async def choose(placeholder: str, value: str):
        for frame in page.frames:
            field = frame.locator(f'input[placeholder="{placeholder}"]').first
            if not await field.is_visible():
                continue
            if placeholder == "排除标签":
                await field.fill(value)
                await field.press("Enter")
                await page.wait_for_timeout(300)
                return
            await field.click()
            await page.wait_for_timeout(500)
            for option_frame in page.frames:
                option = option_frame.get_by_text(value, exact=True).last
                if await option.is_visible():
                    await option.click()
                    await page.wait_for_timeout(300)
                    return
            raise RuntimeError(f"filter option not found: {placeholder}={value}")
        raise RuntimeError(f"filter field not found: {placeholder}")

    await choose("仓位类型", location_type)
    await choose("仓库区域", warehouse_zone)
    await choose("排除标签", exclude_tag)

    # Click 查询
    clicked = False
    for frame in page.frames:
        query_btn = frame.locator('button:has-text("查询")').first
        if await query_btn.is_visible():
            await query_btn.click()
            clicked = True
            break
    if not clicked:
        raise RuntimeError("query button not found")
    await page.wait_for_timeout(2500)

    log.info("JST: filters applied")


async def _start_export(page) -> bool:
    """Click the PackItem toolbar's exact 导出/导出商品 action."""
    for frame in page.frames:
        if not re.search(r"/Wms/(?:Pack/PackItem|Inventory/WarehouseInventory)", frame.url, re.I):
            continue
        info = await frame.evaluate(r"""() => {
          const all = [...document.querySelectorAll('i[title="导入导出"],i[title="导出"],i.icondaorudaochu')];
          const el = all.find(node => { const r=node.getBoundingClientRect(); return r.width>0 && r.height>0; });
          if (!el) return null;
          let parent=el, popoverId=null;
          for (let i=0;i<6 && parent;i++) { parent=parent.parentElement; if (parent) { popoverId=parent.getAttribute('aria-describedby'); if (popoverId) break; } }
          el.click(); return {popoverId};
        }""")
        if not info:
            continue
        await page.wait_for_timeout(500)
        popover_id = info.get("popoverId")
        if popover_id:
            clicked = await frame.evaluate(r"""pid => {
              const pop=document.getElementById(pid); if (!pop) return false;
              const option=[...pop.querySelectorAll('.right-btn-box')].find(el => /^\s*导出(?:商品)?\s*$/.test(el.textContent));
              if (!option) return false; option.click(); return true;
            }""", popover_id)
            if clicked:
                return True
        option = frame.get_by_text(re.compile(r"^(?:导出|导出商品)$")).first
        if await option.is_visible():
            await option.evaluate("el => el.click()")
            return True
    return False


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def _parse_inventory_xlsx(raw: bytes) -> list[dict]:
    """Parse JST inventory Excel → list of {part_code, part_name, on_hand_qty, location, image_url}."""
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    # JST exports an incorrect A1:A1 worksheet dimension even when thousands
    # of rows are present. Force openpyxl to scan the real sheet XML.
    ws.reset_dimensions()

    # Detect header row
    headers = []
    data_start = 2
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or all(v is None for v in row):
            continue
        vals = [str(v).strip() if v is not None else "" for v in row]
        lower_vals = [v.lower() for v in vals]
        if any("sku" in v or "编码" in v or "配件" in v for v in lower_vals):
            headers = vals
            data_start = i + 1
            break

    if not headers:
        headers = ["", "SKU", "名称", "数量", "仓位", "图片"]
        data_start = 2

    code_col = next((i for i, h in enumerate(headers) if "sku" in h.lower() or "编码" in h or "配件编码" in h), 1)
    name_col = next((i for i, h in enumerate(headers) if "名称" in h or "品名" in h), 2)
    qty_col  = next((i for i, h in enumerate(headers) if "数量" in h or "库存" in h), 3)
    loc_col  = next((i for i, h in enumerate(headers) if "仓位" in h or "位置" in h), 4)
    img_col  = next((i for i, h in enumerate(headers) if "图片" in h or "image" in h.lower()), -1)

    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or all(v is None for v in row):
            continue
        vals = [str(v).strip() if v is not None else "" for v in row]
        def g(col: int) -> str:
            return vals[col] if col < len(vals) else ""
        code = g(code_col)
        if not code:
            continue
        try:
            qty = int(float(g(qty_col) or 0))
        except ValueError:
            qty = 0
        rows.append({
            "part_code":   code,
            "part_name":   g(name_col),
            "on_hand_qty": qty,
            "location":    g(loc_col),
            "image_url":   g(img_col) if img_col >= 0 else "",
        })

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Main: fetch from all 3 accounts, merge, dedupe
# ---------------------------------------------------------------------------

async def fetch_pick_locations() -> list[dict]:
    """Launch browser, log in to 66/88/99, apply 拣货仓位 filters, export, parse.

    Returns merged list of inventory rows across all accounts.
    """
    download_dir = Path("/tmp/jst_downloads").resolve()
    download_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[dict] = []

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=HEADLESS,
            executable_path=CHROME,
            args=["--no-sandbox", "--disable-setuid-sandbox", "--disable-dev-shm-usage"],
        )
        ctx = await browser.new_context(
            accept_downloads=True,
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        )
        page = await ctx.new_page()

        for acct in JST_ACCOUNTS:
            account_id = acct["id"]
            email = acct["email"]

            log.info("JST %s: starting", account_id)
            download_path = None

            def make_handler(path_ref):
                async def handle_download(download):
                    path_ref[0] = download.path()
                    log.info("JST %s download: %s", account_id, path_ref[0])
                return handle_download

            try:
                await _login(page, email, account_id)
                await _open_warehouse_inventory(page)
                await _apply_filters(page)
                await page.screenshot(path=f"/tmp/jst-all-sku-{account_id}-filtered.png", full_page=True)

                # Click 导出 and capture the exact download.
                async with page.expect_download(timeout=DOWNLOAD_WAIT * 1000) as pending:
                    if not await _start_export(page):
                        raise RuntimeError("export control not found")
                download = await pending.value
                saved = Path("/tmp") / f"jst-all-sku-{account_id}.xlsx"
                await download.save_as(saved)
                raw = saved.read_bytes()
                rows = _parse_inventory_xlsx(raw)
                log.info("JST %s: parsed %d rows", account_id, len(rows))
                all_rows.extend(rows)

            except Exception as exc:
                log.exception("JST %s error: %s", account_id, exc)
                continue

        await ctx.close()

    if not all_rows:
        raise RuntimeError("No data fetched from any JST account (66/88/99)")

    log.info("Total rows fetched: %d (across all accounts)", len(all_rows))
    return all_rows


def merge_catalogue(rows: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for row in rows:
        sku = row["part_code"].strip()
        location = row["location"].strip()
        if not sku or not location:
            continue
        current = merged.setdefault(sku, {"sku": sku, "location": [], "image_url": row.get("image_url", "")})
        if location not in current["location"]:
            current["location"].append(location)
        if not current["image_url"] and row.get("image_url"):
            current["image_url"] = row["image_url"]
    return [{**row, "location": "、".join(row["location"])} for row in merged.values()]


async def upload_catalogue(rows: list[dict]) -> dict:
    import urllib.request
    auth = _keychain_json("jstreturn-admin-login")
    if not auth:
        raise RuntimeError("protected jstreturn admin credential unavailable")
    base = os.environ.get("JSTRETURN_BASE_URL", "https://jstreturn.onrender.com")
    login_body = json.dumps({"name": auth["name"], "token": auth["token"]}).encode()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor())
    req = urllib.request.Request(base + "/api/auth/login", data=login_body, headers={"Content-Type": "application/json"})
    with opener.open(req, timeout=30):
        pass
    body = json.dumps({"rows": rows}, ensure_ascii=False).encode()
    req = urllib.request.Request(base + "/api/spx/all-sku/import", data=body, headers={"Content-Type": "application/json"})
    with opener.open(req, timeout=120) as response:
        return json.loads(response.read())


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    log.info("Starting JST pick-position fetcher (66/88/99)")

    if "--healthcheck" in sys.argv:
        required = [Path(CHROME).exists(), all(_keychain_json(f"jstreturn-jsterp-{i}") or _keychain_json(f"openclaw-jst-{i}-runtime") for i in ("66", "88", "99")), bool(_keychain_json("jstreturn-admin-login"))]
        print(json.dumps({"ok": all(required), "checks": len(required)}))
        raise SystemExit(0 if all(required) else 1)

    rows = asyncio.run(fetch_pick_locations())

    # Dedupe by part_code (keep first seen)
    merged = merge_catalogue(rows)

    log.info("Merged unique SKUs: %d", len(merged))
    for r in merged[:5]:
        print(r["sku"], r["location"])

    # Save to temp JSON for downstream processing
    out = Path("/tmp/jst_pick_locations.json")
    out.write_text(json.dumps(merged, ensure_ascii=False, indent=2))
    log.info("Saved to %s", out)
    if "--no-upload" not in sys.argv:
        result = asyncio.run(upload_catalogue(merged))
        log.info("Uploaded catalogue rows: %s", result.get("count"))
