"""SPX shipment processing: upload, lookup, pick-list.

Workflow:
  1. Upload SPX Excel → rows parsed & stored in spx_shipments
  2. Scan Tracking No. → return SKUs + our location + employee-entered location
  3. Admin print pick-list → filter by date, show all AWBs ready for picking
  4. JST fetcher for new-pick locations (拣货仓位 / 主仓 / exclude 配件)
"""
from __future__ import annotations

import io
import json
import logging
import re
from datetime import date, datetime, time, timedelta
from typing import Optional
from zoneinfo import ZoneInfo

import openpyxl
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse
from pydantic import BaseModel

from app.auth import require_role
from app.db import pool

router = APIRouter(prefix="/api/spx", tags=["spx"])
log = logging.getLogger("jstreturn.spx")
KLT = ZoneInfo("Asia/Kuala_Lumpur")

# ---------------------------------------------------------------------------
# SKU parsing helpers
# ---------------------------------------------------------------------------

# Matches e.g.  ST5822PK-001  → base=ST5822PK, seq=1
#              ABC-100       → base=ABC, seq=100
SEQ_PATTERN = re.compile(r"^(.+?)-(\d{1,3})$")


def parse_sku(raw: str) -> tuple[str, int]:
    """Return (sku, qty) from a raw cell.

    Rules:
    - 'SKU*2'  → qty=2
    - 'SKU'    → qty=1
    - Trailing whitespace and surrounding newlines stripped.
    """
    raw = raw.strip()
    m = re.match(r"^(.+?)\s*\*\s*(\d+)$", raw)
    if m:
        return m.group(1).strip(), int(m.group(2))
    return raw, 1


def base_sku(sku: str) -> str:
    """Strip the -001 … -100 suffix; return the base SKU."""
    m = SEQ_PATTERN.match(sku)
    if m:
        return m.group(1)
    return sku


def decode_items_json(value) -> list[dict]:
    """Normalize asyncpg JSONB output (string by default) to a list."""
    if value is None:
        return []
    if isinstance(value, str):
        value = json.loads(value)
    if not isinstance(value, list):
        raise ValueError("items_json must be a list")
    return value


def parse_create_time(value: str) -> Optional[datetime]:
    """Parse common SPX timestamps and interpret naive values as Malaysia time."""
    raw = str(value or "").strip()
    if not raw:
        return None
    parsed = None
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00").replace(" ", "T"))
    except ValueError:
        for fmt in (
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
        ):
            try:
                parsed = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
    if parsed is None:
        return None
    return parsed.replace(tzinfo=KLT) if parsed.tzinfo is None else parsed


async def resolve_location(conn, sku: str) -> Optional[str]:
    """Look up location for a SKU.

    1. Try exact SKU
    2. Try base SKU (sku without -001/-002/... suffix)
    3. Return None if neither found.
    """
    row = await conn.fetchrow(
        "SELECT location FROM inventory_snapshot WHERE part_code = $1 AND on_hand_qty > 0 LIMIT 1",
        sku,
    )
    if row:
        return row["location"] or ""

    base = base_sku(sku)
    if base != sku:
        row = await conn.fetchrow(
            "SELECT location FROM inventory_snapshot WHERE part_code = $1 AND on_hand_qty > 0 LIMIT 1",
            base,
        )
        if row:
            return row["location"] or ""

    return None


async def resolve_parts_sku_details(conn, sku: str) -> Optional[dict]:
    """Return parts-inventory image/location, trying exact then base SKU."""
    row = await conn.fetchrow(
        """SELECT location, COALESCE(image_url, '') AS image_url
           FROM inventory_snapshot
           WHERE UPPER(TRIM(part_code)) = UPPER(TRIM($1))
           LIMIT 1""",
        sku,
    )
    if row:
        return dict(row)

    base = base_sku(sku)
    if base != sku:
        row = await conn.fetchrow(
            """SELECT location, COALESCE(image_url, '') AS image_url
               FROM inventory_snapshot
               WHERE UPPER(TRIM(part_code)) = UPPER(TRIM($1))
               LIMIT 1""",
            base,
        )
        if row:
            return dict(row)
    return None


async def resolve_all_sku_details(conn, sku: str) -> Optional[dict]:
    """Return the new-goods image/location, trying exact then base SKU."""
    row = await conn.fetchrow(
        "SELECT location, image_url FROM spx_all_sku_inventory WHERE UPPER(TRIM(sku)) = UPPER(TRIM($1))",
        sku,
    )
    if row:
        return dict(row)

    base = base_sku(sku)
    if base != sku:
        row = await conn.fetchrow(
            "SELECT location, image_url FROM spx_all_sku_inventory WHERE UPPER(TRIM(sku)) = UPPER(TRIM($1))",
            base,
        )
        if row:
            return dict(row)
    return None


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

# Column indices (0-based) in the SPX report
COL_REPORT_TIME   = 0   # 'Report Download Time'
COL_TRACKING      = 0   # 'Tracking No.'
COL_ITEM_IN_PARCEL = 28 # 'Item in Parcel'
COL_NO_OF_ITEMS   = 29  # 'No. of item in Parcel'
COL_ITEM_LIST     = 30  # 'Item List'
COL_CREATE_TIME   = 4   # 'Create Time'


def _col_idx(ws, name: str) -> int:
    """Find column index by header name (case-insensitive)."""
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if str(cell.value or "").strip().lower() == name.lower():
                return cell.column - 1
    raise ValueError(f"Column not found: {name!r}")


def _parse_item_in_parcel(raw: str) -> list[tuple[str, int, str]]:
    """Parse 'Item in Parcel' cell → list of (sku, qty, employee_location).

    Format per line in cell: 'SKU\n位置'
    Multiple lines separated by \n.
    Returns [(sku, qty, employee_location_or_empty), ...]
    """
    lines = [ln.strip() for ln in (raw or "").split("\n") if ln.strip()]
    result = []
    i = 0
    while i < len(lines):
        sku, qty = parse_sku(lines[i])
        loc = ""
        # If next non-empty line doesn't look like a SKU (no letters, just a location pattern),
        # treat it as the employee-entered location.
        if i + 1 < len(lines):
            next_line = lines[i + 1]
            # A location line typically has digits/hyphens but no typical SKU letters
            # or is just 6+ chars with hyphens. Be permissive: if it's not a SKU with qty marker.
            if not re.search(r"[A-Za-z]{2,}", next_line) and re.search(r"[\d\w]{4,}", next_line):
                loc = next_line
                i += 1
        result.append((sku, qty, loc))
        i += 1
    return result


def _col(ws, name: str) -> int:
    try:
        return _col_idx(ws, name)
    except ValueError:
        return -1


def _normalize_header(value) -> str:
    """Normalize exported Excel headers for reliable matching."""
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", str(value or "").strip().lower())


SPX_HEADER_ALIASES = {
    "tracking": {
        "trackingno", "trackingnumber", "trackingid", "awb", "awbno",
        "运单号", "物流单号", "追踪号",
    },
    "parcel": {
        "iteminparcel", "itemsinparcel", "itemlist", "sku", "skulist",
        "包裹商品", "包裹内商品", "商品sku", "商品",
    },
    "create_time": {
        "createtime", "createdtime", "ordertime", "下单时间", "创建时间",
    },
}


def _find_spx_header(ws) -> tuple[int, dict[str, Optional[int]]] | None:
    """Find a usable SPX header row, tolerating titles and header variants."""
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1
    ):
        normalized = [_normalize_header(value) for value in row]
        found = {
            key: next((i for i, value in enumerate(normalized) if value in aliases), None)
            for key, aliases in SPX_HEADER_ALIASES.items()
        }
        if found["tracking"] is not None and found["parcel"] is not None:
            return row_number, found
    return None


def _find_fixed_layout_start(ws) -> Optional[int]:
    """Find the first data row in SPX's stable 31-column export layout.

    Some regional SPX exports translate or omit the header labels while
    retaining the documented column positions (tracking=1, create time=5,
    item in parcel=29). Use that layout only when both required data cells
    are populated, so unrelated workbooks are not silently accepted.
    """
    if ws.max_column is not None and ws.max_column <= COL_ITEM_IN_PARCEL:
        return None
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=100, values_only=True), start=1
    ):
        if len(row) <= COL_ITEM_IN_PARCEL:
            continue
        tracking = str(row[COL_TRACKING] or "").strip()
        parcel = str(row[COL_ITEM_IN_PARCEL] or "").strip()
        normalized_tracking = _normalize_header(tracking)
        if (
            tracking
            and parcel
            and bool(re.search(r"\d", tracking))
            and not bool(re.search(r"[\u4e00-\u9fff]", tracking))
            and normalized_tracking not in SPX_HEADER_ALIASES["tracking"]
            and normalized_tracking not in {"reportdownloadtime", "报表下载时间"}
        ):
            return row_number
    return None


def parse_spx_xlsx(raw: bytes) -> list[dict]:
    """Parse SPX Excel workbook → list of shipment dicts.

    Returns [ {
        tracking_no, create_time, items: [(sku, qty, employee_location), ...]
    }, ... ]
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        # Some SPX exports contain a broken worksheet dimension of A1:A1 even
        # though the XML contains the full table. In read-only mode openpyxl
        # trusts that marker and otherwise exposes only the first cell.
        for sheet in wb.worksheets:
            if sheet.max_row == 1 and sheet.max_column == 1:
                sheet.reset_dimensions()
        selected = next(
            ((ws, match) for ws in wb.worksheets if (match := _find_spx_header(ws)) is not None),
            None,
        )
        if selected is not None:
            ws, (header_row_number, col_map) = selected
            data_start_row = header_row_number + 1
            tracking_col = col_map["tracking"]
            parcel_col = col_map["parcel"]
            time_col = col_map.get("create_time")
        else:
            fixed = next(
                ((ws, start) for ws in wb.worksheets if (start := _find_fixed_layout_start(ws)) is not None),
                None,
            )
            if fixed is None:
                raise ValueError(
                    "required SPX columns not found: need Tracking No. and Item in Parcel/SKU"
                )
            ws, data_start_row = fixed
            tracking_col = COL_TRACKING
            parcel_col = COL_ITEM_IN_PARCEL
            time_col = COL_CREATE_TIME

        rows = []
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not row or all(v is None for v in row):
                continue
            tracking = str(row[tracking_col] or "").strip()
            if not tracking:
                continue
            create_time = ""
            if time_col is not None and time_col < len(row) and row[time_col]:
                create_time = str(row[time_col])
            items = []
            if parcel_col < len(row) and row[parcel_col]:
                items = _parse_item_in_parcel(str(row[parcel_col]))
            if items:
                rows.append({
                    "tracking_no": tracking,
                    "create_time": create_time,
                    "items": items,
                })
        return rows
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------

class ShipmentItemOut(BaseModel):
    sku: str
    qty: int
    employee_location: str
    our_location: Optional[str] = None  # None = not in stock
    image_url: str = ""


class ShipmentOut(BaseModel):
    tracking_no: str
    create_time: str
    items: list[ShipmentItemOut]


class UploadResult(BaseModel):
    total_rows: int
    saved_rows: int


class PickListItem(BaseModel):
    tracking_no: str
    create_time: str
    sku: str
    qty: int
    our_location: Optional[str] = None
    employee_location: str = ""


class PickListOut(BaseModel):
    date: str
    items: list[PickListItem]
    total: int


class AllSkuRow(BaseModel):
    sku: str
    location: str
    image_url: str = ""


class AllSkuImport(BaseModel):
    rows: list[AllSkuRow]


# ---------------------------------------------------------------------------
# DB: create table if not exists
# ---------------------------------------------------------------------------

SPX_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS public.spx_shipments (
    id              BIGSERIAL PRIMARY KEY,
    tracking_no     TEXT NOT NULL,
    create_time     TIMESTAMPTZ,
    items_json      JSONB NOT NULL,   -- [{sku, qty, employee_location}]
    uploaded_at     TIMESTAMPTZ DEFAULT NOW(),
    UNIQUE (tracking_no)
);
CREATE INDEX IF NOT EXISTS idx_spx_tracking ON public.spx_shipments (tracking_no);
CREATE INDEX IF NOT EXISTS idx_spx_uploaded ON public.spx_shipments (uploaded_at);
"""

ALL_SKU_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS public.spx_all_sku_inventory (
    sku         TEXT PRIMARY KEY,
    location    TEXT NOT NULL,
    image_url   TEXT NOT NULL DEFAULT '',
    updated_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_spx_all_sku_location ON public.spx_all_sku_inventory (location);
"""


async def ensure_spx_table():
    async with pool().acquire() as conn:
        await conn.execute(SPX_TABLE_SQL)


async def ensure_all_sku_table():
    async with pool().acquire() as conn:
        await conn.execute(ALL_SKU_TABLE_SQL)


@router.post("/all-sku/import")
async def import_all_sku(
    payload: AllSkuImport = Body(...),
    user: dict = Depends(require_role("admin")),
):
    """Atomically replace the separate new-goods SKU/location catalogue."""
    normalized = {}
    for row in payload.rows:
        sku = row.sku.strip()
        location = row.location.strip()
        if sku and location:
            normalized[sku] = (location, row.image_url.strip())
    if not normalized:
        raise HTTPException(400, "all-SKU import is empty")
    await ensure_all_sku_table()
    async with pool().acquire() as conn:
        async with conn.transaction():
            await conn.execute("DELETE FROM spx_all_sku_inventory")
            await conn.executemany(
                "INSERT INTO spx_all_sku_inventory (sku, location, image_url) VALUES ($1, $2, $3)",
                [(sku, loc, image) for sku, (loc, image) in normalized.items()],
            )
    return {"ok": True, "count": len(normalized)}


@router.get("/all-sku")
async def list_all_sku(
    q: str = "",
    limit: int = Query(200, ge=1, le=1000),
    user: dict = Depends(require_role("admin", "repair")),
):
    await ensure_all_sku_table()
    term = q.strip()
    async with pool().acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT sku, location, image_url, updated_at
            FROM spx_all_sku_inventory
            WHERE ($1 = '' OR sku ILIKE '%' || $1 || '%' OR location ILIKE '%' || $1 || '%')
            ORDER BY sku
            LIMIT $2
            """,
            term, limit,
        )
        total = await conn.fetchval(
            "SELECT COUNT(*) FROM spx_all_sku_inventory WHERE ($1 = '' OR sku ILIKE '%' || $1 || '%' OR location ILIKE '%' || $1 || '%')",
            term,
        )
    return {"count": int(total or 0), "items": [dict(row) for row in rows]}


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.post("/upload")
async def upload_spx(
    file: UploadFile = File(...),
    user: dict = Depends(require_role("admin", "returns")),
):
    """Upload SPX Excel file.

    Parses Tracking No. + Item in Parcel (SKU + employee location),
    saves to spx_shipments table, and returns a summary.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "must be .xlsx")

    await ensure_spx_table()

    raw = await file.read()
    try:
        rows = parse_spx_xlsx(raw)
    except Exception as exc:
        log.exception("SPX parse error")
        raise HTTPException(400, f"parse error: {exc}") from exc

    if not rows:
        raise HTTPException(400, "no shipment rows found")

    saved = 0
    async with pool().acquire() as conn:
        for row in rows:
            items_json = [
                {"sku": sku, "qty": qty, "employee_location": loc}
                for sku, qty, loc in row["items"]
            ]
            create_ts = None
            if row["create_time"]:
                create_ts = parse_create_time(row["create_time"])
            result = await conn.execute(
                """
                INSERT INTO spx_shipments (tracking_no, create_time, items_json)
                VALUES ($1, $2, $3::jsonb)
                ON CONFLICT (tracking_no) DO UPDATE SET
                    create_time = EXCLUDED.create_time,
                    items_json = EXCLUDED.items_json,
                    uploaded_at = NOW()
                """,
                row["tracking_no"],
                create_ts,
                json.dumps(items_json, ensure_ascii=False),
            )
            if result.startswith("INSERT") or result.startswith("UPDATE"):
                saved += 1

    return UploadResult(total_rows=len(rows), saved_rows=saved)


@router.get("/lookup/{tracking_no}", response_model=ShipmentOut)
async def lookup_tracking(
    tracking_no: str,
    user: dict = Depends(require_role("admin", "repair")),
):
    """Scan a Tracking No. → return SKUs, our location, employee location."""

    await ensure_spx_table()
    await ensure_all_sku_table()
    async with pool().acquire() as conn:
        row = await conn.fetchrow(
            """SELECT tracking_no,
                      COALESCE(create_time, uploaded_at) AS effective_time,
                      items_json
               FROM spx_shipments
               WHERE UPPER(TRIM(tracking_no)) = UPPER(TRIM($1))""",
            tracking_no,
        )

    if not row:
        raise HTTPException(404, f"Tracking {tracking_no!r} not found")

    items_out = []
    async with pool().acquire() as conn:
        for item in decode_items_json(row["items_json"]):
            sku = item.get("sku", "")
            all_sku = await resolve_all_sku_details(conn, sku)
            parts_sku = await resolve_parts_sku_details(conn, sku)
            our_loc = ((all_sku or {}).get("location")
                       or (parts_sku or {}).get("location")
                       or "无库存")
            items_out.append(ShipmentItemOut(
                sku=sku,
                qty=item.get("qty", 1),
                employee_location=item.get("employee_location", ""),
                our_location=our_loc,
                image_url=((all_sku or {}).get("image_url")
                           or (parts_sku or {}).get("image_url")
                           or ""),
            ))

    return ShipmentOut(
        tracking_no=row["tracking_no"],
        create_time=str(row["effective_time"] or ""),
        items=items_out,
    )


@router.get("/pick-list", response_model=PickListOut)
async def pick_list(
    date_str: str = Query(..., description="Date in YYYY-MM-DD"),
    user: dict = Depends(require_role("admin")),
):
    """Print pick-list for all SPX shipments on a given date (YYYY-MM-DD)."""

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "date_str must be YYYY-MM-DD")

    await ensure_spx_table()
    await ensure_all_sku_table()
    start = datetime.combine(target_date, time.min, tzinfo=KLT)
    end = start + timedelta(days=1)

    async with pool().acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT tracking_no,
                   COALESCE(create_time, uploaded_at) AS effective_time,
                   items_json
            FROM spx_shipments
            WHERE uploaded_at >= $1
              AND uploaded_at < $2
            ORDER BY uploaded_at, tracking_no
            """,
            start, end,
        )

    items_out: list[PickListItem] = []
    async with pool().acquire() as conn:
        for row in rows:
            for item in decode_items_json(row["items_json"]):
                sku = item.get("sku", "")
                all_sku = await resolve_all_sku_details(conn, sku)
                parts_sku = await resolve_parts_sku_details(conn, sku)
                our_loc = ((all_sku or {}).get("location")
                           or (parts_sku or {}).get("location")
                           or "无库存")
                items_out.append(PickListItem(
                    tracking_no=row["tracking_no"],
                    create_time=str(row["effective_time"] or ""),
                    sku=sku,
                    qty=item.get("qty", 1),
                    our_location=our_loc,
                    employee_location=item.get("employee_location", ""),
                ))

    return PickListOut(
        date=date_str,
        items=items_out,
        total=len(items_out),
    )
