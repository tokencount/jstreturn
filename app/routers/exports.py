"""Purchase-export: roll up PENDING missing-part quantities per part_code.

The export uses Asia/Kuala_Lumpur business_date semantics. The pure
aggregation helper (`compute_purchase_rows`) is split out from the HTTP
handler so unit tests can hit it without a live database.

ZIP output bundles one xlsx per warehouse prefix (HS / HE / HU), each
derived from an upstream Excel template (`media/inbound/.../poin_1…xlsx`).
Each xlsx has two sheets — Sheet1 (the cloned template header + that
prefix's purchase rows, D:Z left blank) and Sheet2 (verbatim copy of
the upstream configuration table).

When no prefix has any purchase rows, the ZIP contains a single
`empty-YYYY-MM-DD.txt` placeholder instead of empty xlsx files.
"""
from __future__ import annotations

import io
import json
import logging
import zipfile
from collections import defaultdict
from datetime import date as _date, datetime, timezone
from pathlib import Path
from typing import Iterable, Mapping, Optional
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl import Workbook, load_workbook

from app.auth import current_user
from app.matcher import compute_reservation_plan

log = logging.getLogger("jstreturn")

# Asia/Kuala_Lumpur is the project's business-day reference. All
# callers default to "today in that TZ" if no business_date is supplied.
BUSINESS_TZ = ZoneInfo("Asia/Kuala_Lumpur")

# The three warehouses supported in the customer contract.
WAREHOUSE_MAP: dict[str, str] = {
    "HS": "HS168-自营",
    "HE": "HE168-自营仓",
    "HU": "HU168-自营仓",
}

# Path to the upstream Excel template. The template is shipped INSIDE
# the deployed repo at app/templates/purchase_template.xlsx so it is
# always present on Render (`/opt/render/project/src/app/templates/...`)
# without depending on any workspace-level `media/inbound/...` folder.
#
# Resolution:
#   exports.py -> app/routers/ -> app/ -> templates/purchase_template.xlsx
#
# Tests can still monkey-patch TEMPLATE_PATH or pass an explicit
# template_path= into build_purchase_zip().
TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "templates" / "purchase_template.xlsx"
)


def _resolve_template_path() -> Path:
    """Return the on-disk template path, falling back to the legacy
    workspace-level media/inbound location if the in-repo template is
    missing.

    Keeps production deployments self-contained (Render only ships the
    repo, not the workspace `media/` folder) while letting devs that
    keep their template under media/inbound keep working without an
    extra copy step.
    """
    if TEMPLATE_PATH.exists():
        return TEMPLATE_PATH
    # Legacy fallback: media/inbound/<stage-id>/poin_1---...xlsx
    legacy = (
        Path(__file__).resolve().parent.parent.parent.parent
        / "media"
        / "inbound"
        / "openclaw-staged-74fc7f7e-9d66-4109-a8b3-76b76906f50f"
        / "poin_1---4900ea7d-5685-4f68-a71d-a3a55621efb0.xlsx"
    )
    return legacy

router = APIRouter(prefix="/api/exports", tags=["exports"])


# ---------------------------------------------------------------------------
# Pure aggregation logic (no DB, fully unit-testable)
# ---------------------------------------------------------------------------

def _part_prefix(part_code: str) -> str:
    """Return the leading 'XX' (alpha) prefix of a part_code.

    Falls back to '' for codes without two leading letters.
    """
    code = (part_code or "").strip().upper()
    if len(code) >= 2 and code[0].isalpha() and code[1].isalpha():
        return code[:2]
    return ""


def compute_purchase_rows(
    pending_items: Iterable[Mapping],
    inventory_map: Mapping[str, int],
) -> dict:
    """Aggregate part-purchase demand for PENDING items.

    Args:
        pending_items: Iterable of items with `.id` and `.parts`. Each part
            is a mapping with `part_code` and `qty`.
        inventory_map: part_code → on_hand_qty.

    Returns:
        dict with:
          by_code:  dict[part_code, qty_needed]
          by_warehouse: dict["HS"|"HE"|"HU", list[(part_code, qty)]]
          skipped_prefixes: list[str]  sorted unique prefixes seen but not HS/HE/HU
          skipped_part_codes: list[(part_code, qty)]  ordered, sorted by code

    The "qty_needed" for a part_code sums over every PENDING ticket the
    ticket still owes: max(0, part.qty − (on_hand_qty − reserved)), where
    `reserved` is the global reservation consumed by all READY tickets in
    the same plan.
    """
    items = []
    for raw in pending_items:
        item = dict(raw)
        parts = item.get("parts") or []
        item["parts"] = list(parts)
        items.append(item)

    # Run the global reservation plan so we know how much of each part_code
    # is already consumed by READY tickets across the whole system.
    plan = compute_reservation_plan(items, dict(inventory_map))
    reserved_by_part = plan["reserved_by_part"]
    ready_ids = plan["ready_ids"]

    by_code: dict[str, int] = defaultdict(int)
    for raw in items:
        # Only count demand for items that are NOT already reserved by
        # the matcher; READY items are fully covered and their demand
        # is met by the existing reservation, so they do not contribute
        # to purchase demand.
        if int(raw["id"]) in ready_ids:
            continue
        for p in raw.get("parts") or []:
            code = (p.get("part_code") or "").strip()
            if not code:
                continue
            qty = int(p.get("qty") or 0)
            on_hand = int(inventory_map.get(code, 0))
            reserved = int(reserved_by_part.get(code, 0))
            available = max(on_hand - reserved, 0)
            short = max(qty - available, 0)
            if short > 0:
                by_code[code] += short

    by_warehouse: dict[str, list[tuple[str, int]]] = {k: [] for k in WAREHOUSE_MAP}
    skipped_part_codes: list[tuple[str, int]] = []
    skipped_prefixes_set: set[str] = set()

    for code in sorted(by_code):
        prefix = _part_prefix(code)
        qty = by_code[code]
        if prefix in WAREHOUSE_MAP:
            by_warehouse[prefix].append((code, qty))
        else:
            skipped_part_codes.append((code, qty))
            if prefix:
                skipped_prefixes_set.add(prefix)

    return {
        "by_code": dict(by_code),
        "by_warehouse": by_warehouse,
        "skipped_prefixes": sorted(skipped_prefixes_set),
        "skipped_part_codes": skipped_part_codes,
    }


def _copy_cell_style(src_cell, dst_cell) -> None:
    """Copy style from src_cell → dst_cell without altering dst_cell.value."""
    if src_cell.has_style:
        dst_cell.font = src_cell.font.copy()
        dst_cell.fill = src_cell.fill.copy()
        dst_cell.border = src_cell.border.copy()
        dst_cell.alignment = src_cell.alignment.copy()
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = src_cell.protection.copy()


def build_purchase_xlsx(
    prefix: str,
    rows: list[tuple[str, int]],
    business_date: _date,
    template_wb: Workbook,
) -> bytes:
    """Build a single xlsx for one warehouse prefix.

    The output workbook has exactly two sheets:
        Sheet1 — clone of template Sheet1 with that prefix's rows under A/B/C
        Sheet2 — verbatim copy of template Sheet2

    Caller is expected to skip calling this for prefixes with no rows so
    we never produce empty xlsx files. (See ``build_purchase_zip``.)
    """
    template_sheet = template_wb["Sheet1"]
    template_sheet2 = template_wb["Sheet2"]
    warehouse_name = WAREHOUSE_MAP[prefix]
    hs_template_cols = template_sheet.max_column  # 26

    # Clone Sheet1 into a temporary title, then strip the original so the
    # final name is exactly "Sheet1".
    cloned = template_wb.copy_worksheet(template_sheet)
    cloned.title = "__tmp_clone__"

    # Now drop the original Sheet1.
    template_wb.remove(template_sheet)

    # Rename the clone to "Sheet1".
    cloned.title = "Sheet1"

    # Wipe any cloned data rows from row 2 onwards so we don't carry
    # template values into the export.
    for r in range(2, cloned.max_row + 1):
        for c in range(1, hs_template_cols + 1):
            cloned.cell(r, c).value = None

    # Copy Sheet2 verbatim (used by downstream ERP-style importers).
    cloned2 = template_wb.copy_worksheet(template_sheet2)
    cloned2.title = "__tmp_sheet2__"
    template_wb.remove(template_sheet2)
    cloned2.title = "Sheet2"

    # Reorder sheets so Sheet1 comes before Sheet2 regardless of the
    # template's internal order.
    template_wb._sheets = [cloned, cloned2]

    for i, (code, qty) in enumerate(rows):
        row_idx = 2 + i
        src_a = cloned.cell(1, 1)
        dst_a = cloned.cell(row_idx, 1)
        _copy_cell_style(src_a, dst_a)
        dst_a.value = warehouse_name

        src_b = cloned.cell(1, 2)
        dst_b = cloned.cell(row_idx, 2)
        _copy_cell_style(src_b, dst_b)
        dst_b.value = code

        src_c = cloned.cell(1, 3)
        dst_c = cloned.cell(row_idx, 3)
        _copy_cell_style(src_c, dst_c)
        dst_c.value = int(qty)

    buf = io.BytesIO()
    template_wb.save(buf)
    return buf.getvalue()


def build_purchase_zip(
    agg: dict,
    business_date: _date,
    template_path: Optional[Path] = None,
) -> bytes:
    """Build the ZIP archive that bundles per-warehouse xlsx files.

    The ZIP contains, in this order:
      * ``HS-采购配件-YYYY-MM-DD.xlsx`` — only if HS has rows
      * ``HE-采购配件-YYYY-MM-DD.xlsx`` — only if HE has rows
      * ``HU-采购配件-YYYY-MM-DD.xlsx`` — only if HU has rows
      * If no prefix has any rows: a single ``empty-YYYY-MM-DD.txt``
        placeholder so the archive is never literally empty.

    We build each xlsx against a freshly-loaded copy of the template
    workbook so a single template is not mutated across prefixes.
    """
    template_path = template_path or _resolve_template_path()
    if not template_path.exists():
        raise HTTPException(500, f"purchase export template missing at {template_path}")

    by_warehouse = agg.get("by_warehouse") or {}
    has_any = any(by_warehouse.get(p) for p in ("HS", "HE", "HU"))

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if not has_any:
            placeholder = (
                f"{business_date.isoformat()} 无采购需求。\n"
                f"被跳过的 part 数: {len(agg.get('skipped_part_codes') or [])}\n"
                f"被跳过的前缀: {', '.join(agg.get('skipped_prefixes') or []) or '—'}\n"
            )
            zf.writestr(f"empty-{business_date.isoformat()}.txt", placeholder)
        else:
            for prefix in ("HS", "HE", "HU"):
                rows = by_warehouse.get(prefix) or []
                if not rows:
                    continue
                # Fresh template per prefix to avoid openpyxl cross-prefix
                # workbook state leakage.
                wb = load_workbook(template_path)
                xlsx_bytes = build_purchase_xlsx(
                    prefix=prefix,
                    rows=rows,
                    business_date=business_date,
                    template_wb=wb,
                )
                filename = f"{prefix}-采购配件-{business_date.isoformat()}.xlsx"
                zf.writestr(filename, xlsx_bytes)
    return buf.getvalue()


def today_in_business_tz(now: Optional[datetime] = None) -> _date:
    """Today's date in Asia/Kuala_Lumpur."""
    now = now or datetime.now(timezone.utc)
    return now.astimezone(BUSINESS_TZ).date()


def parse_business_date(value: Optional[str]) -> _date:
    """Strict YYYY-MM-DD parser."""
    if not value:
        raise HTTPException(400, "business_date required (YYYY-MM-DD)")
    try:
        return _date.fromisoformat(value)
    except ValueError:
        raise HTTPException(400, "business_date must be YYYY-MM-DD")


def build_summary(agg: dict, business_date: _date) -> dict:
    """Shape the JSON summary returned alongside the ZIP."""
    return {
        "business_date": business_date.isoformat(),
        "row_count": sum(len(v) for v in agg["by_warehouse"].values()),
        "by_warehouse": {k: len(v) for k, v in agg["by_warehouse"].items()},
        "warehouses_included": [
            p for p in ("HS", "HE", "HU") if agg["by_warehouse"].get(p)
        ],
        "by_code": dict(agg["by_code"]),
        "skipped_prefixes": agg["skipped_prefixes"],
        "skipped_count": len(agg["skipped_part_codes"]),
        "skipped_part_codes": [list(t) for t in agg["skipped_part_codes"]],
    }


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------

@router.get("/purchase")
async def export_purchase(
    business_date: Optional[str] = Query(
        None,
        description="业务日期 YYYY-MM-DD，Asia/Kuala_Lumpur。省略则取当天的 KLT 日期。",
    ),
    summary_only: bool = Query(
        False,
        description="true 时只返回 JSON 汇总（不下载 ZIP）；UI 显示跳过提示用。",
    ),
    user: dict = Depends(current_user),
):
    """Roll up PENDING parts for `business_date` into a ZIP of xlsx files.

    Response shape:

    * ``application/json`` when ``summary_only=true`` — UI hint pane.
    * ``application/zip`` otherwise — one xlsx per warehouse prefix
      (HS / HE / HU). When no prefix has rows the ZIP contains a single
      ``empty-YYYY-MM-DD.txt`` placeholder.

    The ZIP entry names are ``HS-采购配件-YYYY-MM-DD.xlsx`` etc.; the
    ZIP itself is named ``采购配件-YYYY-MM-DD.zip``.
    """
    target_date = parse_business_date(business_date) if business_date else today_in_business_tz()
    return await _export_purchase_for_date(target_date, user, summary_only=summary_only)


@router.get("/purchase/preview")
async def preview_purchase(
    business_date: Optional[str] = Query(
        None,
        description="业务日期 YYYY-MM-DD，Asia/Kuala_Lumpur。省略则取当天的 KLT 日期。",
    ),
    user: dict = Depends(current_user),
):
    """JSON-only preview of the purchase export for `business_date`.

    Always returns JSON shape (never a ZIP). The frontend uses this to
    display the "被跳过的 part_code / 前缀 / 仓库名" hint pane before the
    user clicks the actual download button.

    Adds a ``skipped`` field per task spec listing every part_code that
    was not assigned to a warehouse plus its prefix and the warehouse
    name it would have mapped to (empty string when no matching prefix).
    """
    target_date = parse_business_date(business_date) if business_date else today_in_business_tz()
    summary = await _export_purchase_for_date(target_date, user, summary_only=True)
    if isinstance(summary, JSONResponse):
        body = summary.body
        if isinstance(body, (bytes, bytearray)):
            body = body.decode("utf-8")
        try:
            data = json.loads(body)
        except Exception:
            data = {}
    elif isinstance(summary, dict):
        data = summary
    else:
        data = {}
    # Build the skipped shape requested by the task:
    #   skipped: [{ part_code, prefix, warehouse_name, qty }]
    skipped = []
    for code, qty in data.get("skipped_part_codes") or []:
        prefix = _part_prefix(code)
        skipped.append({
            "part_code": code,
            "prefix": prefix,
            "warehouse_name": WAREHOUSE_MAP.get(prefix, ""),
            "qty": int(qty),
        })
    data["skipped"] = skipped
    data["skipped_count"] = len(skipped)
    data["skipped_prefixes"] = sorted({s["prefix"] for s in skipped if s["prefix"]})
    return JSONResponse(data)


async def _export_purchase_for_date(target_date: _date, user: dict, summary_only: bool = False):
    """Internal: pull aggregate + build ZIP / JSON summary."""
    from app.db import pool

    async with pool().acquire() as conn:
        item_rows = await conn.fetch(
            """
            SELECT
                di.id, di.pallet_no, di.sku, di.qty, di.status, di.created_at,
                COALESCE(json_agg(json_build_object(
                    'part_code', dp.part_code,
                    'part_name', dp.part_name,
                    'qty', dp.qty
                ) ORDER BY dp.id) FILTER (WHERE dp.id IS NOT NULL), '[]'::json) AS parts
            FROM defective_items di
            LEFT JOIN defective_parts dp ON dp.defective_id = di.id
            WHERE di.status = 'PENDING' AND di.business_date = $1
            GROUP BY di.id
            ORDER BY di.id
            """,
            target_date,
        )
        inventory_rows = await conn.fetch(
            "SELECT part_code, on_hand_qty FROM inventory_snapshot"
        )

    items = []
    for r in item_rows:
        item = dict(r)
        parts = item.get("parts")
        if isinstance(parts, str):
            parts = json.loads(parts)
        item["parts"] = list(parts or [])
        if isinstance(item.get("created_at"), datetime):
            item["created_at"] = item["created_at"]
        items.append(item)

    inventory_map = {r["part_code"]: int(r["on_hand_qty"] or 0) for r in inventory_rows}
    agg = compute_purchase_rows(items, inventory_map)
    summary = build_summary(agg, target_date)

    if summary_only:
        return JSONResponse(summary)

    zip_bytes = build_purchase_zip(agg, target_date)
    filename = f"采购配件-{target_date.isoformat()}.zip"
    # RFC 6266: ``filename*=UTF-8''<percent-encoded>`` for non-ASCII filenames.
    # Starlette encodes headers as latin-1 so we percent-encode the UTF-8 bytes
    # for the wire-format ``filename*`` parameter.
    from urllib.parse import quote
    ascii_fallback = "purchase_parts-" + target_date.isoformat() + ".zip"
    filename_star = quote(filename, safe="")
    headers = {
        "Content-Disposition": (
            f"attachment; filename=\"{ascii_fallback}\"; "
            f"filename*=UTF-8''{filename_star}"
        ),
        "X-Export-Skipped-Prefixes": ",".join(agg["skipped_prefixes"] or []) or "",
        "X-Export-Skipped-Count": str(len(agg["skipped_part_codes"] or [])),
        "X-Export-Business-Date": target_date.isoformat(),
        "X-Export-Total-Parts": str(len(agg["by_code"])),
        "X-Export-Row-Count": str(summary["row_count"]),
        "X-Export-Warehouses-Included": ",".join(
            p for p in ("HS", "HE", "HU") if agg["by_warehouse"].get(p)
        ),
        # Echo the summary as JSON in a header for clients that can't
        # read the body when downloading an attachment.
        "X-Export-Summary": json.dumps(summary, ensure_ascii=True),
    }

    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers=headers,
    )
