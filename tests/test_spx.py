import asyncio
import hashlib
import inspect
import io
import json
import unittest
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

from app.routers import spx


class SpxParserTests(unittest.TestCase):
    def _workbook(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        stream = io.BytesIO()
        wb.save(stream)
        wb.close()
        return stream.getvalue()

    def test_parser_finds_header_and_reads_rows_before_closing_workbook(self):
        raw = self._workbook([
            ["SPX export"],
            ["Tracking No.", "Create Time", "Item in Parcel"],
            ["SPX001", "2026-09-01 05:00:00", "ABC-001*2"],
        ])
        self.assertEqual(
            spx.parse_spx_xlsx(raw),
            [{
                "tracking_no": "SPX001",
                "create_time": "2026-09-01 05:00:00",
                "items": [("ABC-001", 2, "")],
            }],
        )

    def test_parser_rejects_workbook_without_required_columns(self):
        raw = self._workbook([["wrong", "columns"], ["a", "b"]])
        with self.assertRaisesRegex(ValueError, "required SPX columns"):
            spx.parse_spx_xlsx(raw)

    def test_parser_accepts_common_header_variants_and_whitespace(self):
        raw = self._workbook([
            ["Tracking\nNumber", "Created Time", "SKU List"],
            ["SPX002", "2026-09-01 14:00:00", "ABC*2"],
        ])
        self.assertEqual(spx.parse_spx_xlsx(raw)[0]["items"], [("ABC", 2, "")])

    def test_parser_finds_headers_below_row_ten(self):
        rows = [[f"report title {i}"] for i in range(12)]
        rows.extend([
            ["AWB No", "创建时间", "商品 SKU"],
            ["SPX003", "2026-09-01 14:00:00", "XYZ"],
        ])
        self.assertEqual(spx.parse_spx_xlsx(self._workbook(rows))[0]["tracking_no"], "SPX003")

    def test_parser_finds_data_sheet_when_first_sheet_is_cover(self):
        wb = openpyxl.Workbook()
        wb.active.append(["SPX cover"])
        ws = wb.create_sheet("Data")
        ws.append(["运单号", "创建时间", "包裹内商品"])
        ws.append(["SPX004", "2026-09-01 14:00:00", "SKU9*3"])
        stream = io.BytesIO()
        wb.save(stream)
        wb.close()
        parsed = spx.parse_spx_xlsx(stream.getvalue())
        self.assertEqual(parsed[0]["items"], [("SKU9", 3, "")])

    def test_decode_items_json_accepts_asyncpg_string(self):
        self.assertEqual(
            spx.decode_items_json('[{"sku":"ABC","qty":2}]'),
            [{"sku": "ABC", "qty": 2}],
        )

    def test_parse_common_spx_day_first_timestamp_as_malaysia_time(self):
        parsed = spx.parse_create_time("01/09/2026 05:20:30")
        self.assertEqual(parsed, datetime(2026, 9, 1, 5, 20, 30, tzinfo=ZoneInfo("Asia/Kuala_Lumpur")))

    def test_invalid_create_time_returns_none_for_uploaded_at_fallback(self):
        self.assertIsNone(spx.parse_create_time("not-a-date"))


class FakeConnection:
    def __init__(self, rows):
        self.rows = list(rows)
        self.queries = []

    async def fetchrow(self, query, sku):
        self.queries.append(sku)
        return self.rows.pop(0) if self.rows else None


class SpxLocationTests(unittest.IsolatedAsyncioTestCase):
    async def test_exact_location_is_awaited_and_returned(self):
        conn = FakeConnection([{"location": "A-01"}])
        self.assertEqual(await spx.resolve_location(conn, "ABC-001"), "A-01")
        self.assertEqual(conn.queries, ["ABC-001"])

    async def test_base_sku_fallback(self):
        conn = FakeConnection([None, {"location": "B-02"}])
        self.assertEqual(await spx.resolve_location(conn, "ABC-001"), "B-02")
        self.assertEqual(conn.queries, ["ABC-001", "ABC"])

    async def test_all_sku_details_exact_match(self):
        conn = FakeConnection([{"location": "P-01", "image_url": "https://img/abc.jpg"}])
        self.assertEqual(
            await spx.resolve_all_sku_details(conn, "ABC-001"),
            {"location": "P-01", "image_url": "https://img/abc.jpg"},
        )
        self.assertEqual(conn.queries, ["ABC-001"])

    async def test_all_sku_details_base_fallback(self):
        conn = FakeConnection([None, {"location": "P-02", "image_url": "https://img/base.jpg"}])
        self.assertEqual(
            await spx.resolve_all_sku_details(conn, "ABC-001"),
            {"location": "P-02", "image_url": "https://img/base.jpg"},
        )
        self.assertEqual(conn.queries, ["ABC-001", "ABC"])

    async def test_parts_sku_details_exact_match(self):
        conn = FakeConnection([{"location": "PART-01", "image_url": "https://img/part.jpg"}])
        self.assertEqual(
            await spx.resolve_parts_sku_details(conn, "HS-PART-001"),
            {"location": "PART-01", "image_url": "https://img/part.jpg"},
        )
        self.assertEqual(conn.queries, ["HS-PART-001"])

    async def test_parts_sku_details_base_fallback(self):
        conn = FakeConnection([None, {"location": "PART-02", "image_url": "https://img/base-part.jpg"}])
        self.assertEqual(
            await spx.resolve_parts_sku_details(conn, "HS-PART-001"),
            {"location": "PART-02", "image_url": "https://img/base-part.jpg"},
        )
        self.assertEqual(conn.queries, ["HS-PART-001", "HS-PART"])


class SpxContractTests(unittest.TestCase):
    def test_lookup_route_matches_visible_roles(self):
        source = inspect.getsource(spx.lookup_tracking)
        self.assertNotIn("returns", source)

    def test_upload_route_allows_admin_and_returns(self):
        source = inspect.getsource(spx.upload_spx)
        self.assertIn('require_role("admin", "returns")', source)
        self.assertIn('endswith(".xlsx")', source)
        self.assertNotIn('".xls"', source)

    def test_jsonb_payload_is_serialized(self):
        source = Path(spx.__file__).read_text(encoding="utf-8")
        self.assertIn("json.dumps(items_json, ensure_ascii=False)", source)

    def test_all_database_routes_initialize_the_spx_table(self):
        self.assertIn("await ensure_spx_table()", inspect.getsource(spx.upload_spx))
        self.assertIn("await ensure_spx_table()", inspect.getsource(spx.lookup_tracking))
        self.assertIn("await ensure_spx_table()", inspect.getsource(spx.pick_list))

    def test_lookup_normalizes_tracking_and_decodes_jsonb(self):
        source = inspect.getsource(spx.lookup_tracking)
        self.assertIn("UPPER(TRIM(tracking_no))", source)
        self.assertIn("UPPER(TRIM($1))", source)
        self.assertIn("decode_items_json", source)
        self.assertIn("resolve_all_sku_details", source)
        self.assertIn("resolve_parts_sku_details", source)
        self.assertIn("image_url", source)
        self.assertIn('or "无库存"', source)

    def test_pick_list_filters_the_uploaded_batch_and_decodes_jsonb(self):
        source = inspect.getsource(spx.pick_list)
        self.assertIn("WHERE uploaded_at >= $1", source)
        self.assertIn("AND uploaded_at < $2", source)
        self.assertIn("decode_items_json", source)
        self.assertIn("resolve_all_sku_details", source)
        self.assertIn("resolve_parts_sku_details", source)
        self.assertIn('or "无库存"', source)

    def test_all_sku_catalogue_is_separate_from_parts_inventory(self):
        source = Path(spx.__file__).read_text(encoding="utf-8")
        self.assertIn("spx_all_sku_inventory", source)
        import_source = inspect.getsource(spx.import_all_sku)
        self.assertNotIn("inventory_snapshot", import_source)
        self.assertIn('require_role("admin")', import_source)

    def test_ui_contains_buttons_sections_and_http_error_handling(self):
        html = (Path(__file__).parents[1] / "app/templates/index.html").read_text(encoding="utf-8")
        for marker in ("tab==='spx'", "spxView", "发货上传", "运单查询", "拣货单"):
            self.assertIn(marker, html)
        self.assertEqual(html.count("@click=\"goTab('spx')\""), 1)
        self.assertIn('<section x-show="tab===\'spx\'" x-cloak class="repair-allow">', html)
        self.assertIn('class="nav-btn repair-keep"', html)
        self.assertNotIn("goTab('spx-upload')", html)
        self.assertNotIn("goTab('spx-lookup')", html)
        self.assertNotIn("goTab('spx-pick')", html)
        self.assertIn("['admin','repair','returns'].includes(user.role)", html)
        self.assertIn("['admin','returns'].includes(user.role)", html)
        self.assertIn("this.spxUploadResult = r.ok ? data", html)
        self.assertIn("spxUploadResult: null", html)
        self.assertIn("ev.target.value = ''", html)
        self.assertIn("this.spxPickResult = r.ok ? data", html)
        self.assertIn("查询全部", html)
        self.assertIn("spxSelectedTrackings", html)
        self.assertIn("generateSpxPickSummary()", html)
        self.assertIn("拣货单 · SKU 汇总", html)
        self.assertIn("数量总和", html)
        self.assertIn("<th>仓位</th>", html)
        self.assertNotIn("我们的仓位", html)
        self.assertNotIn("员工仓位", html)
        self.assertIn("item.our_location || item.employee_location || '—'", html)
        self.assertIn("<th>图片</th>", html)
        self.assertIn("x-show=\"item.image_url\"", html)
        self.assertIn("const location = item.our_location || item.employee_location", html)
        self.assertIn("removeSpxPickSku(row.sku)", html)
        self.assertIn("removeSpxPickSku(sku)", html)
        self.assertIn("row.sku !== sku", html)
        self.assertIn("打印 A4", html)
        self.assertIn(':disabled="!spxPickGenerated"', html)
        self.assertIn("@page { size: A4 portrait", html)
        self.assertIn("spx-pick-print-area", html)
        self.assertIn("@click=\"window.print()\"", html)
        self.assertIn("class=\"no-print\">操作", html)
        self.assertIn("All SKU 库存", html)
        self.assertIn("照片</th><th>SKU</th><th>仓位", html)
        self.assertIn("/api/spx/all-sku", html)


if __name__ == "__main__":
    unittest.main()
