"""Purchase-export ZIP tests.

These tests pin the public contract of ``GET /api/exports/purchase``:

  * returns ``application/zip`` with ``Content-Disposition``
  * each per-prefix xlsx has exactly Sheet1 + Sheet2
  * Sheet1 header row matches the upstream template (all 26 columns)
  * Sheet1 column widths match the template (within +/- 0.5)
  * Sheet1 A1 has non-default font / fill / alignment
  * Sheet2 is byte-equivalent to the template's Sheet2
  * rows below the header only ever fill A/B/C; D-Z stay empty
  * quantity aggregation is per part_code (not per ticket SKU)
  * ``/preview`` endpoint returns JSON with a `skipped` list
  * empty => ZIP contains ``empty-YYYY-MM-DD.txt``
  * time zone is ``Asia/Kuala_Lumpur`` (UTC+8)

Aggregation is unit-tested separately (compute_purchase_rows).
"""
import io
import json
import unittest
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import app.routers.exports as exports_mod
from app.auth import current_user
import app.db as db_mod  # noqa: F401  # pool is imported lazily inside the handler

TEMPLATE_PATH = exports_mod.TEMPLATE_PATH
BUSINESS_TZ = exports_mod.BUSINESS_TZ


# ---------------------------------------------------------------------------
# Mock asyncpg pool + DB primitives
# ---------------------------------------------------------------------------

class _FakeConn:
    def __init__(self):
        self.fetchrow = AsyncMock()
        self.fetch = AsyncMock(return_value=[])
        self.execute = AsyncMock()
        self.executemany = AsyncMock()

    def transaction(self):
        outer = self
        class _TxCM:
            async def __aenter__(self_inner):
                return outer
            async def __aexit__(self_inner, *exc):
                return None
        return _TxCM()

    async def __aenter__(self):
        return self

    async def __aexit__(self, *exc):
        return None


def make_pool_with(conn):
    pool = MagicMock()
    class _PoolCM:
        async def __aenter__(self_inner):
            return conn
        async def __aexit__(self_inner, *exc):
            return None
    pool.acquire = MagicMock(return_value=_PoolCM())
    return pool


def _build_app(role="returns"):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient
    app = FastAPI()
    app.include_router(exports_mod.router)
    fake_user = {"id": 1, "role": role, "active": True, "name": "test"}
    app.dependency_overrides[current_user] = lambda: fake_user
    return app, fake_user


def _seed_db(conn, defective_items, inventory_map):
    item_rows = []
    for it in defective_items:
        item_rows.append({**it, "parts": json.dumps(it["parts"])})
    inv_rows = [
        {"part_code": code, "on_hand_qty": qty}
        for code, qty in inventory_map.items()
    ]
    conn.fetch = AsyncMock(side_effect=[item_rows, inv_rows])
    return conn


SAMPLE_ITEMS = [
    {
        "id": 1, "pallet_no": "PLT-001", "sku": "SKU-1", "qty": 1, "status": "PENDING",
        "created_at": datetime(2026, 8, 14, 1, 0, 0, tzinfo=timezone.utc),
        "parts": [
            {"part_code": "HS-AAA-001", "qty": 3},
            {"part_code": "HE-BBB-002", "qty": 2},
        ],
    },
    {
        "id": 2, "pallet_no": "PLT-002", "sku": "SKU-2", "qty": 1, "status": "PENDING",
        "created_at": datetime(2026, 8, 14, 2, 0, 0, tzinfo=timezone.utc),
        "parts": [
            {"part_code": "HS-AAA-001", "qty": 1},
            {"part_code": "HU-CCC-003", "qty": 5},
            {"part_code": "UN-999-001", "qty": 7},
        ],
    },
]

SAMPLE_INVENTORY = {
    "HS-AAA-001": 10,
    "HE-BBB-002": 0,
    "HU-CCC-003": 1,
}


# ---------------------------------------------------------------------------
# Pure aggregation helper tests
# ---------------------------------------------------------------------------

class ComputePurchaseRowsTests(unittest.TestCase):
    def test_empty_inputs(self):
        out = exports_mod.compute_purchase_rows([], {})
        self.assertEqual(out["by_code"], {})
        self.assertEqual(out["by_warehouse"], {"HS": [], "HE": [], "HU": []})
        self.assertEqual(out["skipped_part_codes"], [])

    def test_aggregation_per_part_code_not_per_ticket(self):
        items = [
            {"id": 1, "pallet_no": "P1", "parts": [
                {"part_code": "HS-AAA-001", "qty": 3},
                {"part_code": "HS-AAA-001", "qty": 2},
            ]},
            {"id": 2, "pallet_no": "P2", "parts": [
                {"part_code": "HS-AAA-001", "qty": 5},
            ]},
        ]
        inv = {"HS-AAA-001": 0}
        out = exports_mod.compute_purchase_rows(items, inv)
        self.assertEqual(out["by_code"]["HS-AAA-001"], 10)
        self.assertEqual(out["by_warehouse"]["HS"], [("HS-AAA-001", 10)])

    def test_short_capped_at_zero(self):
        items = [{"id": 1, "pallet_no": "P", "parts": [{"part_code": "HS-AAA-001", "qty": 3}]}]
        inv = {"HS-AAA-001": 100}
        out = exports_mod.compute_purchase_rows(items, inv)
        self.assertNotIn("HS-AAA-001", out["by_code"])

    def test_skipped_prefixes_listed(self):
        items = [{"id": 1, "pallet_no": "P", "parts": [
            {"part_code": "UN-999", "qty": 1},
            {"part_code": "ZZ-999", "qty": 2},
        ]}]
        inv = {"UN-999": 0, "ZZ-999": 0}
        out = exports_mod.compute_purchase_rows(items, inv)
        self.assertEqual(set(out["skipped_prefixes"]), {"UN", "ZZ"})
        self.assertEqual(len(out["skipped_part_codes"]), 2)

    def test_reserved_by_part_deducts_from_available(self):
        items = [
            {"id": 1, "pallet_no": "P1", "parts": [{"part_code": "HS-AAA-001", "qty": 8}]},
            {"id": 2, "pallet_no": "P2", "parts": [{"part_code": "HS-AAA-001", "qty": 5}]},
        ]
        inv = {"HS-AAA-001": 10}
        out = exports_mod.compute_purchase_rows(items, inv)
        # READY item 1 reserves 8; PENDING item 2 needs 5 with available 2 = short 3
        self.assertEqual(out["by_code"]["HS-AAA-001"], 3)

    def test_part_prefix_helper(self):
        from app.routers.exports import _part_prefix
        self.assertEqual(_part_prefix("HS-AAA-001"), "HS")
        self.assertEqual(_part_prefix("HE"), "HE")
        self.assertEqual(_part_prefix(""), "")
        self.assertEqual(_part_prefix("1ABC"), "")


# ---------------------------------------------------------------------------
# ZIP endpoint tests
# ---------------------------------------------------------------------------

class PurchaseExportZipTests(unittest.TestCase):
    def _call(self, role="returns", business_date="2026-08-14", summary=False):
        self.app, self.user = _build_app(role)
        self.conn = _FakeConn()
        _seed_db(self.conn, SAMPLE_ITEMS, SAMPLE_INVENTORY)
        self._pool_patch = patch.object(db_mod, "pool", lambda: make_pool_with(self.conn))
        self._pool_patch.start()
        from fastapi.testclient import TestClient
        self.client = TestClient(self.app)
        params = {"business_date": business_date}
        if summary:
            params["summary_only"] = "true"
        return self.client.get("/api/exports/purchase", params=params)

    def tearDown(self):
        try:
            self._pool_patch.stop()
        except Exception:
            pass

    def test_returns_zip_with_correct_content_type(self):
        r = self._call()
        self.assertEqual(r.status_code, 200, r.text)
        self.assertEqual(r.headers["content-type"], "application/zip")
        cd = r.headers.get("content-disposition", "")
        self.assertIn("attachment", cd)
        # UTF-8 percent-encoded filename OR ASCII fallback
        self.assertTrue(
            "purchase_parts-2026-08-14.zip" in cd
            or "采购配件-2026-08-14.zip" in cd
        )

    def test_zip_contains_only_warehouses_with_rows(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        names = z.namelist()
        # HS-AAA-001 has stock 10 > 3+1=4 → no purchase
        self.assertNotIn("HS-采购配件-2026-08-14.xlsx", names)
        self.assertIn("HE-采购配件-2026-08-14.xlsx", names)
        self.assertIn("HU-采购配件-2026-08-14.xlsx", names)

    def test_zip_summary_headers_present(self):
        r = self._call()
        self.assertEqual(r.headers["x-export-business-date"], "2026-08-14")
        self.assertIn("UN", r.headers["x-export-skipped-prefixes"].split(","))
        wh = r.headers["x-export-warehouses-included"].split(",")
        self.assertIn("HE", wh)
        self.assertIn("HU", wh)
        self.assertNotIn("HS", wh)
        summary = json.loads(r.headers["x-export-summary"])
        self.assertEqual(summary["business_date"], "2026-08-14")
        # HE-BBB-002 short=2, HU-CCC-003 short=4 → 2 rows total
        self.assertEqual(summary["row_count"], 2)
        self.assertIn("by_warehouse", summary)

    def test_xlsx_sheet_names(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HE-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        self.assertEqual(wb.sheetnames, ["Sheet1", "Sheet2"])

    def test_xlsx_sheet1_header_matches_template(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HE-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Sheet1"]
        self.assertEqual(ws.max_column, 26)
        tpl_wb = load_workbook(TEMPLATE_PATH)
        tpl = tpl_wb["Sheet1"]
        for c in range(1, 27):
            self.assertEqual(ws.cell(1, c).value, tpl.cell(1, c).value)

    def test_xlsx_sheet1_columns_width_match_template(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HE-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Sheet1"]
        tpl_wb = load_workbook(TEMPLATE_PATH)
        tpl = tpl_wb["Sheet1"]
        for letter in ("B", "F", "K", "L", "N", "O", "P", "X"):
            tw = tpl.column_dimensions[letter].width
            gw = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
            self.assertIsNotNone(gw)
            self.assertAlmostEqual(float(gw), float(tw), delta=0.5)

    def test_xlsx_sheet1_a1_has_non_default_style(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HE-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Sheet1"]
        a1 = ws["A1"]
        self.assertIsNotNone(a1.font)
        self.assertIsNotNone(a1.font.name)
        self.assertNotEqual(a1.font.name, "Calibri")
        self.assertIsNotNone(a1.fill)
        self.assertIsNotNone(a1.alignment)

    def test_xlsx_sheet2_matches_template(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HE-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws2 = wb["Sheet2"]
        tpl_wb = load_workbook(TEMPLATE_PATH)
        tpl2 = tpl_wb["Sheet2"]
        self.assertEqual(ws2.max_row, tpl2.max_row)
        self.assertEqual(ws2.max_column, tpl2.max_column)
        for r_ in range(1, tpl2.max_row + 1):
            for c_ in range(1, tpl2.max_column + 1):
                self.assertEqual(
                    ws2.cell(r_, c_).value,
                    tpl2.cell(r_, c_).value,
                    f"Sheet2 {r_},{c_} mismatch",
                )

    def test_xlsx_data_rows_only_in_a_b_c(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HU-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Sheet1"]
        for row in range(2, ws.max_row + 1):
            for c in range(4, 27):
                self.assertIsNone(
                    ws.cell(row, c).value,
                    f"row {row} col {c} should be empty",
                )

    def test_xlsx_data_rows_correct_warehouse_and_code(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        xlsx = z.read("HE-采购配件-2026-08-14.xlsx")
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb["Sheet1"]
        self.assertEqual(ws.cell(2, 1).value, "HE168-自营仓")
        self.assertEqual(ws.cell(2, 2).value, "HE-BBB-002")
        self.assertEqual(ws.cell(2, 3).value, 2)

    def test_quantity_aggregation_per_part_code(self):
        r = self._call()
        z = zipfile.ZipFile(io.BytesIO(r.content))
        names = z.namelist()
        self.assertNotIn("HS-采购配件-2026-08-14.xlsx", names)
        summary = json.loads(r.headers["x-export-summary"])
        self.assertNotIn("HS-AAA-001", summary["by_code"])
        self.assertEqual(summary["by_code"]["HE-BBB-002"], 2)
        self.assertEqual(summary["by_code"]["HU-CCC-003"], 4)

    def test_summary_only_returns_json(self):
        r = self._call(summary=True)
        self.assertEqual(r.status_code, 200, r.text)
        self.assertTrue(r.headers["content-type"].startswith("application/json"))
        body = r.json()
        self.assertEqual(body["business_date"], "2026-08-14")
        self.assertEqual(body["by_code"]["HE-BBB-002"], 2)
        self.assertEqual(body["by_code"]["HU-CCC-003"], 4)
        self.assertNotIn("HS-AAA-001", body["by_code"])

    def test_default_business_date_uses_klt_tz(self):
        from zoneinfo import ZoneInfo as _ZI
        KLT = _ZI("Asia/Kuala_Lumpur")
        now_utc = datetime.now(timezone.utc)
        klt_today = now_utc.astimezone(KLT).date()
        items = [
            {
                "id": 7, "pallet_no": "PLT-K", "sku": "SKU-K", "qty": 1, "status": "PENDING",
                "created_at": now_utc,
                "parts": [{"part_code": f"HE-{klt_today.isoformat().replace('-', '')}", "qty": 1}],
            },
        ]
        inv = {f"HE-{klt_today.isoformat().replace('-', '')}": 0}
        self.app, self.user = _build_app("returns")
        self.conn = _FakeConn()
        _seed_db(self.conn, items, inv)
        self._pool_patch = patch.object(db_mod, "pool", lambda: make_pool_with(self.conn))
        self._pool_patch.start()
        from fastapi.testclient import TestClient
        c = TestClient(self.app)
        r = c.get("/api/exports/purchase")
        self.assertEqual(r.status_code, 200, r.text)
        self.assertEqual(r.headers["x-export-business-date"], klt_today.isoformat())

    def test_klt_time_zone_boundary(self):
        target = date(2026, 8, 14)
        items = [
            {
                "id": 1, "pallet_no": "P", "sku": "S", "qty": 1, "status": "PENDING",
                "created_at": datetime(2026, 8, 14, 0, 30, tzinfo=timezone.utc),
                "parts": [{"part_code": "HE-D1", "qty": 1}],
            },
        ]
        inv = {"HE-D1": 0}
        self.app, self.user = _build_app("returns")
        self.conn = _FakeConn()
        rows = [{**items[0], "parts": json.dumps(items[0]["parts"])}]
        inv_rows = [{"part_code": "HE-D1", "on_hand_qty": 0}]
        self.conn.fetch = AsyncMock(side_effect=[rows, inv_rows])
        self._pool_patch = patch.object(db_mod, "pool", lambda: make_pool_with(self.conn))
        self._pool_patch.start()
        from fastapi.testclient import TestClient
        c = TestClient(self.app)
        r = c.get("/api/exports/purchase", params={"business_date": target.isoformat()})
        self.assertEqual(r.status_code, 200, r.text)
        summary = json.loads(r.headers["x-export-summary"])
        self.assertIn("HE-D1", summary["by_code"])


# ---------------------------------------------------------------------------
# Empty data test
# ---------------------------------------------------------------------------

class PurchaseExportEmptyTests(unittest.TestCase):
    def _call(self):
        self.app, self.user = _build_app("returns")
        self.conn = _FakeConn()
        self.conn.fetch = AsyncMock(side_effect=[[], []])
        self._pool_patch = patch.object(db_mod, "pool", lambda: make_pool_with(self.conn))
        self._pool_patch.start()
        from fastapi.testclient import TestClient
        self.client = TestClient(self.app)
        return self.client.get("/api/exports/purchase", params={"business_date": "2026-08-14"})

    def tearDown(self):
        try:
            self._pool_patch.stop()
        except Exception:
            pass

    def test_empty_zip_contains_placeholder(self):
        r = self._call()
        self.assertEqual(r.status_code, 200, r.text)
        self.assertEqual(r.headers["content-type"], "application/zip")
        z = zipfile.ZipFile(io.BytesIO(r.content))
        names = z.namelist()
        self.assertEqual(len(names), 1)
        self.assertEqual(names[0], "empty-2026-08-14.txt")
        body = z.read(names[0]).decode("utf-8")
        self.assertIn("2026-08-14", body)
        self.assertIn("无采购需求", body)


# ---------------------------------------------------------------------------
# Preview endpoint tests
# ---------------------------------------------------------------------------

class PurchaseExportPreviewTests(unittest.TestCase):
    def _call(self):
        self.app, self.user = _build_app("returns")
        self.conn = _FakeConn()
        _seed_db(self.conn, SAMPLE_ITEMS, SAMPLE_INVENTORY)
        self._pool_patch = patch.object(db_mod, "pool", lambda: make_pool_with(self.conn))
        self._pool_patch.start()
        from fastapi.testclient import TestClient
        self.client = TestClient(self.app)
        return self.client.get("/api/exports/purchase/preview", params={"business_date": "2026-08-14"})

    def tearDown(self):
        try:
            self._pool_patch.stop()
        except Exception:
            pass

    def test_preview_returns_json_skipped_section(self):
        r = self._call()
        self.assertEqual(r.status_code, 200, r.text)
        self.assertTrue(r.headers["content-type"].startswith("application/json"))
        body = r.json()
        self.assertIn("skipped", body)
        self.assertIsInstance(body["skipped"], list)
        un = next((s for s in body["skipped"] if s["part_code"] == "UN-999-001"), None)
        self.assertIsNotNone(un)
        self.assertEqual(un["prefix"], "UN")
        self.assertEqual(un["warehouse_name"], "")
        self.assertEqual(un["qty"], 7)
        self.assertIn("UN", body["skipped_prefixes"])

    def test_preview_never_returns_zip(self):
        r = self._call()
        self.assertNotIn("application/zip", r.headers["content-type"])


class PurchaseExportParamTests(unittest.TestCase):
    def test_invalid_business_date_returns_400(self):
        self.app, self.user = _build_app("returns")
        self.conn = _FakeConn()
        self._pool_patch = patch.object(db_mod, "pool", lambda: make_pool_with(self.conn))
        self._pool_patch.start()
        from fastapi.testclient import TestClient
        c = TestClient(self.app)
        r = c.get("/api/exports/purchase", params={"business_date": "not-a-date"})
        self.assertEqual(r.status_code, 400, r.text)


# ---------------------------------------------------------------------------
# Regression tests for the deployed-path fix (template must live INSIDE
# the deployed repo at app/templates/purchase_template.xlsx so Render
# can find it without needing any workspace-level media/ folder).
# ---------------------------------------------------------------------------

class TemplatePathRegressionTests(unittest.TestCase):
    def test_template_path_lives_under_app_templates(self):
        """TEMPLATE_PATH must point inside app/templates so it ships with
        the deployed repo on Render (project root = /opt/render/project/src).
        """
        p = exports_mod.TEMPLATE_PATH
        self.assertTrue(p.is_absolute(), f"template path should be absolute, got {p}")
        parts = p.parts
        # .../app/templates/purchase_template.xlsx
        self.assertIn("app", parts)
        self.assertIn("templates", parts)
        self.assertEqual(p.name, "purchase_template.xlsx")
        # The path must NOT escape the app-src repo root.
        self.assertNotIn("media", parts, "template should not depend on media/ folder")
        self.assertNotIn("inbound", parts)

    def test_template_file_exists_at_deployed_path(self):
        """The xlsx file must exist at TEMPLATE_PATH on disk so Render
        can load it after the repo is cloned to /opt/render/project/src.
        """
        self.assertTrue(
            exports_mod.TEMPLATE_PATH.exists(),
            f"template missing at {exports_mod.TEMPLATE_PATH}",
        )

    def test_resolve_template_path_returns_existing_file(self):
        p = exports_mod._resolve_template_path()
        self.assertTrue(p.exists(), f"_resolve_template_path returned missing path {p}")
        # Should prefer the in-repo path over the legacy media/ one.
        self.assertEqual(p, exports_mod.TEMPLATE_PATH)

    def test_template_path_relative_to_exports_module(self):
        """Even if exports.py is moved deeper in the tree, the template
        path should still resolve to <exports_dir>/../templates/... (i.e.
        a sibling-of-app module path), not to some absolute workspace path.
        """
        from app.routers import exports as exports_mod_check
        expected = (
            Path(exports_mod_check.__file__).resolve().parent.parent
            / "templates"
            / "purchase_template.xlsx"
        )
        self.assertEqual(exports_mod_check.TEMPLATE_PATH, expected)

    def test_empty_prefix_file_included_when_only_some_warehouses_match(self):
        """Verify the ZIP contains one xlsx per warehouse prefix that has
        rows, with empty-prefix files emitted as required. When a prefix
        has zero rows it is skipped (not emitted as empty xlsx)."""
        items = [
            {
                "id": 1, "pallet_no": "P", "sku": "S", "qty": 1, "status": "PENDING",
                "created_at": datetime(2026, 8, 14, 1, 0, 0, tzinfo=timezone.utc),
                "parts": [{"part_code": "HE-ONLY-001", "qty": 5}],
            },
        ]
        inv = {"HE-ONLY-001": 0}
        app_, _user = _build_app("returns")
        conn = _FakeConn()
        _seed_db(conn, items, inv)
        with patch.object(db_mod, "pool", lambda: make_pool_with(conn)):
            from fastapi.testclient import TestClient
            c = TestClient(app_)
            r = c.get("/api/exports/purchase", params={"business_date": "2026-08-14"})
        self.assertEqual(r.status_code, 200, r.text)
        z = zipfile.ZipFile(io.BytesIO(r.content))
        names = sorted(z.namelist())
        # Only HE has rows → only HE-*.xlsx is in the ZIP; HS/HU omitted.
        self.assertEqual(names, ["HE-采购配件-2026-08-14.xlsx"])


class TemplateGuardRegressionTests(unittest.TestCase):
    """Static checks that the Alpine template guards null access on
    purchaseSummary so the page does not throw
    `Cannot read properties of null (reading 'skipped_count')` before
    the user has run an export.
    """

    TEMPLATE_FILE = Path(__file__).resolve().parent.parent / "app" / "templates" / "index.html"

    def _load_template(self) -> str:
        self.assertTrue(self.TEMPLATE_FILE.exists(), f"missing {self.TEMPLATE_FILE}")
        return self.TEMPLATE_FILE.read_text(encoding="utf-8")

    def test_purchase_skipped_count_guarded_with_null_check(self):
        html = self._load_template()
        # The original bug: `x-show="purchaseSummary.skipped_count > 0"`
        # threw when purchaseSummary was null. Must now be guarded.
        self.assertIn("purchaseSummary && purchaseSummary.skipped_count", html)

    def test_purchase_business_date_guarded(self):
        html = self._load_template()
        self.assertIn("(purchaseSummary && purchaseSummary.business_date)", html)

    def test_purchase_by_warehouse_guarded(self):
        html = self._load_template()
        self.assertIn(
            "(purchaseSummary && purchaseSummary.by_warehouse && purchaseSummary.by_warehouse.HS)",
            html,
        )

    def test_purchase_skipped_part_codes_guarded(self):
        html = self._load_template()
        self.assertIn(
            "(purchaseSummary && purchaseSummary.skipped_part_codes)",
            html,
        )

    def test_repair_ready_inner_xif_replaced_with_xshow(self):
        """The READY tab inner x-if inside an x-for caused Alpine
        `Cannot read properties of null (reading 'after')` page errors.
        The empty-state must now use x-show (no <template x-if> nested
        inside the rc-parts x-for).
        """
        import re
        html = self._load_template()
        # Find the rc-part-lines block (everything between <div class="rc-part-lines">
        # and its closing </div>).
        m = re.search(
            r'<div class="rc-part-lines">\s*(.*?)\s*</div>\s*</article>',
            html,
            re.DOTALL,
        )
        self.assertIsNotNone(m, "rc-part-lines block not found")
        rc_parts = m.group(1)
        # No nested <template x-if> allowed inside the rc-parts x-for.
        self.assertNotIn("<template x-if", rc_parts, "x-if still nested inside rc-parts x-for")
        # x-show empty-state must exist.
        self.assertIn('x-show="!((it.parts || []).filter(', rc_parts)

    def test_mobile_parts_xif_replaced_with_xshow(self):
        """Same fix for the mobile parts list x-for."""
        import re
        html = self._load_template()
        # The mobile parts-list uses <template x-for="p in (it.parts || [])">
        # Match the whole block from the template opener up to the </ul> closing
        # the parts-list (so we capture both the x-for template and the
        # sibling x-show empty-state).
        m = re.search(
            r'<template x-for="p in \(it\.parts \|\| \[\]\)" :key="\'mp-\' \+ it\.id \+ \'-\' \+ p\.part_code">.*?</ul>',
            html,
            re.DOTALL,
        )
        self.assertIsNotNone(m, "mobile parts x-for block not found")
        body = m.group(0)
        # No nested <template x-if> inside the x-for template.
        # (The x-for template ends at </template>; check only inside it.)
        inside_template = re.search(
            r'<template x-for="p in \(it\.parts \|\| \[\]\)" :key="\'mp-\' \+ it\.id \+ \'-\' \+ p\.part_code">(.*?)</template>',
            body,
            re.DOTALL,
        )
        self.assertIsNotNone(inside_template, "inner template not found")
        self.assertNotIn(
            "<template x-if",
            inside_template.group(1),
            "x-if still nested inside mobile parts x-for",
        )
        # x-show empty-state must exist as a sibling.
        self.assertIn('x-show="!(it.parts && it.parts.length)"', body)


if __name__ == "__main__":
    unittest.main()
